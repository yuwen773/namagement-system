import csv
import io
import json
from datetime import datetime, time
from decimal import Decimal, InvalidOperation

from django.db.models import OuterRef, Q, Subquery
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import extend_schema, extend_schema_view, inline_serializer
from rest_framework import filters, mixins, serializers, viewsets
from rest_framework.decorators import action, parser_classes
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.devices.models import Device, EnergyType
from apps.energy.models import EnergyData, EnergyStatistics
from apps.energy.serializers import (
    EnergyDataBatchSerializer,
    EnergyDataSerializer,
    EnergyStatisticsSerializer,
)
from energy_monitoring.permissions import IsAdminOrReadOnly


EnergyBatchImportResponseSerializer = inline_serializer(
    name="EnergyBatchImportResponse",
    fields={
        "imported_count": serializers.IntegerField(),
        "submitted_count": serializers.IntegerField(),
        "skipped_count": serializers.IntegerField(),
        "source_format": serializers.CharField(),
    },
)
EnergyLatestItemSerializer = inline_serializer(
    name="EnergyLatestItem",
    fields={
        "device": serializers.IntegerField(),
        "device_code": serializers.CharField(),
        "device_name": serializers.CharField(),
        "energy_type": serializers.CharField(),
        "timestamp": serializers.DateTimeField(),
        "value": serializers.DecimalField(max_digits=20, decimal_places=6),
        "voltage": serializers.DecimalField(max_digits=20, decimal_places=6, allow_null=True),
        "current": serializers.DecimalField(max_digits=20, decimal_places=6, allow_null=True),
        "power": serializers.DecimalField(max_digits=20, decimal_places=6, allow_null=True),
        "flow_rate": serializers.DecimalField(max_digits=20, decimal_places=6, allow_null=True),
    },
)


def _as_aware_datetime(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        raw = str(value).strip()
        if not raw:
            return None
        dt = parse_datetime(raw)
        if dt is None:
            parsed_date = parse_date(raw)
            if parsed_date is None:
                return None
            dt = datetime.combine(parsed_date, time.min)
    if timezone.is_naive(dt):
        return timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def _parse_decimal(value, field_name, required=False):
    if value in (None, ""):
        if required:
            raise serializers.ValidationError({field_name: f"{field_name} 不能为空。"})
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        raise serializers.ValidationError({field_name: f"{field_name} 不是有效数值。"})


def _normalize_row(row):
    normalized = {}
    for key, value in row.items():
        if key is None:
            continue
        normalized[str(key).strip().lower()] = value
    return normalized


def _pick_value(row, keys):
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return None


@extend_schema_view(
    list=extend_schema(summary="查询能耗原始数据列表"),
    create=extend_schema(summary="新增单条能耗数据", request=EnergyDataSerializer, responses={201: EnergyDataSerializer}),
)
class EnergyDataViewSet(mixins.ListModelMixin, mixins.CreateModelMixin, viewsets.GenericViewSet):
    """能耗原始数据录入、查询、导入导出接口。"""

    queryset = EnergyData.objects.select_related("device", "energy_type").all()
    serializer_class = EnergyDataSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["device__device_id", "device__name"]
    ordering_fields = ["id", "timestamp", "value", "created_at"]
    ordering = ["-timestamp", "-id"]

    def get_queryset(self):
        queryset = super().get_queryset()

        device_id = self.request.query_params.get("device_id")
        if device_id:
            if str(device_id).isdigit():
                queryset = queryset.filter(device_id=int(device_id))
            else:
                queryset = queryset.filter(device__device_id=device_id)

        room_id = self.request.query_params.get("room_id")
        if room_id:
            queryset = queryset.filter(device__room_id=room_id)

        energy_type = self.request.query_params.get("energy_type")
        if energy_type:
            if str(energy_type).isdigit():
                queryset = queryset.filter(energy_type_id=int(energy_type))
            else:
                queryset = queryset.filter(energy_type__code__iexact=energy_type)

        start_date = self.request.query_params.get("start_date")
        if start_date:
            start_dt = _as_aware_datetime(start_date)
            if start_dt is not None:
                queryset = queryset.filter(timestamp__gte=start_dt)

        end_date = self.request.query_params.get("end_date")
        if end_date:
            end_dt = _as_aware_datetime(end_date)
            if end_dt is not None:
                parsed_date = parse_date(str(end_date))
                if parsed_date is not None and "T" not in str(end_date):
                    end_dt = timezone.make_aware(
                        datetime.combine(parsed_date, time.max),
                        timezone.get_current_timezone(),
                    )
                queryset = queryset.filter(timestamp__lte=end_dt)

        return queryset

    def perform_create(self, serializer):
        instance = serializer.save()
        Device.objects.filter(pk=instance.device_id).filter(
            Q(last_data_time__isnull=True) | Q(last_data_time__lt=instance.timestamp)
        ).update(last_data_time=instance.timestamp)

    @action(
        detail=False,
        methods=["post"],
        url_path="batch-import",
        parser_classes=[MultiPartParser, FormParser, JSONParser],
    )
    @extend_schema(
        summary="批量导入能耗数据",
        request=EnergyDataBatchSerializer,
        responses={200: EnergyBatchImportResponseSerializer},
    )
    def batch_import(self, request):
        """批量导入 CSV/Excel/JSON 能耗数据。"""
        serializer = EnergyDataBatchSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        upload_file = serializer.validated_data.get("file")
        records = serializer.validated_data.get("records")
        explicit_format = serializer.validated_data.get("format")

        if upload_file is not None:
            source_format = explicit_format or self._guess_file_format(upload_file.name)
            rows = self._parse_uploaded_rows(upload_file, source_format)
        else:
            source_format = explicit_format or "json"
            if isinstance(records, list):
                rows = records
            elif isinstance(records, dict):
                rows = records.get("records", [])
            else:
                raise serializers.ValidationError({"records": "records 必须是数组或包含 records 数组的对象。"})
            if not isinstance(rows, list):
                raise serializers.ValidationError({"records": "records 必须是数组。"})

        instances = self._build_energy_data_instances(rows)
        before_count = EnergyData.objects.count()
        EnergyData.objects.bulk_create(
            instances,
            batch_size=1000,
            ignore_conflicts=True,
        )
        after_count = EnergyData.objects.count()
        imported_count = max(after_count - before_count, 0)

        latest_timestamp_by_device = {}
        for item in instances:
            latest = latest_timestamp_by_device.get(item.device_id)
            if latest is None or latest < item.timestamp:
                latest_timestamp_by_device[item.device_id] = item.timestamp

        for device_pk, ts in latest_timestamp_by_device.items():
            Device.objects.filter(pk=device_pk).filter(
                Q(last_data_time__isnull=True) | Q(last_data_time__lt=ts)
            ).update(last_data_time=ts)

        return Response(
            {
                "imported_count": imported_count,
                "submitted_count": len(instances),
                "skipped_count": max(len(instances) - imported_count, 0),
                "source_format": source_format,
            }
        )

    @extend_schema(
        summary="获取各设备最新能耗数据",
        responses={200: OpenApiTypes.OBJECT},
    )
    @action(detail=False, methods=["get"], url_path="latest")
    def latest(self, request):
        """查询设备维度的最新采样值。"""
        devices = Device.objects.select_related("energy_type").all().order_by("id")

        device_id = request.query_params.get("device_id")
        if device_id:
            if str(device_id).isdigit():
                devices = devices.filter(pk=int(device_id))
            else:
                devices = devices.filter(device_id=device_id)

        room_id = request.query_params.get("room_id")
        if room_id:
            devices = devices.filter(room_id=room_id)

        energy_type = request.query_params.get("energy_type")
        if energy_type:
            if str(energy_type).isdigit():
                devices = devices.filter(energy_type_id=int(energy_type))
            else:
                devices = devices.filter(energy_type__code__iexact=energy_type)

        latest_source = EnergyData.objects.filter(device_id=OuterRef("pk"))

        start_date = request.query_params.get("start_date")
        if start_date:
            start_dt = _as_aware_datetime(start_date)
            if start_dt is not None:
                latest_source = latest_source.filter(timestamp__gte=start_dt)

        end_date = request.query_params.get("end_date")
        if end_date:
            end_dt = _as_aware_datetime(end_date)
            if end_dt is not None:
                parsed_end_date = parse_date(str(end_date))
                if parsed_end_date is not None and "T" not in str(end_date):
                    end_dt = timezone.make_aware(
                        datetime.combine(parsed_end_date, time.max),
                        timezone.get_current_timezone(),
                    )
                latest_source = latest_source.filter(timestamp__lte=end_dt)

        latest_source = latest_source.order_by("-timestamp", "-id")

        devices = devices.annotate(
            latest_timestamp=Subquery(latest_source.values("timestamp")[:1]),
            latest_value=Subquery(latest_source.values("value")[:1]),
            latest_voltage=Subquery(latest_source.values("voltage")[:1]),
            latest_current=Subquery(latest_source.values("current")[:1]),
            latest_power=Subquery(latest_source.values("power")[:1]),
            latest_flow_rate=Subquery(latest_source.values("flow_rate")[:1]),
        ).filter(latest_timestamp__isnull=False)

        payload = [
            {
                "device": device.id,
                "device_code": device.device_id,
                "device_name": device.name,
                "energy_type": device.energy_type.code,
                "timestamp": device.latest_timestamp,
                "value": device.latest_value,
                "voltage": device.latest_voltage,
                "current": device.latest_current,
                "power": device.latest_power,
                "flow_rate": device.latest_flow_rate,
            }
            for device in devices
        ]
        return Response(payload)

    @extend_schema(
        summary="导出能耗数据",
        description="支持 `excel` 和 `pdf` 两种格式。",
        responses={
            (200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): OpenApiTypes.BINARY,
            (200, "application/pdf"): OpenApiTypes.BINARY,
        },
    )
    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """按指定格式导出当前筛选数据。"""
        # 优先使用 file_format，兼容 format 参数
        export_format = request.query_params.get("file_format") or request.query_params.get("format", "excel")
        export_format = export_format.lower()
        queryset = self.get_queryset().select_related("device", "energy_type").order_by("timestamp", "id")
        data = list(queryset)

        if export_format == "excel":
            return self._export_excel(data)
        if export_format == "pdf":
            return self._export_pdf(data)
        raise serializers.ValidationError({"format": "仅支持 excel 或 pdf。"})

    def _guess_file_format(self, file_name):
        lower_name = file_name.lower()
        if lower_name.endswith(".csv"):
            return "csv"
        if lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
            return "excel"
        if lower_name.endswith(".json"):
            return "json"
        raise serializers.ValidationError({"file": "无法识别文件格式，请显式传 format。"})

    def _parse_uploaded_rows(self, upload_file, source_format):
        if source_format == "csv":
            content = upload_file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            return list(reader)

        if source_format == "json":
            content = upload_file.read().decode("utf-8-sig")
            try:
                payload = json.loads(content)
            except json.JSONDecodeError as exc:
                raise serializers.ValidationError({"file": "JSON 文件解析失败。"}) from exc
            if isinstance(payload, list):
                return payload
            if isinstance(payload, dict) and isinstance(payload.get("records"), list):
                return payload["records"]
            raise serializers.ValidationError({"file": "JSON 文件格式错误，需为数组或包含 records 数组。"})

        if source_format == "excel":
            try:
                import openpyxl
            except ImportError as exc:
                raise serializers.ValidationError({"file": "缺少 openpyxl 依赖，无法解析 Excel。"}) from exc

            workbook = openpyxl.load_workbook(upload_file, read_only=True, data_only=True)
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if headers is None:
                return []
            normalized_headers = [str(value).strip() if value is not None else "" for value in headers]
            rows = []
            for row_values in rows_iter:
                if row_values is None:
                    continue
                row_dict = {}
                empty_row = True
                for idx, header in enumerate(normalized_headers):
                    if not header:
                        continue
                    cell_value = row_values[idx] if idx < len(row_values) else None
                    if cell_value not in (None, ""):
                        empty_row = False
                    row_dict[header] = cell_value
                if not empty_row:
                    rows.append(row_dict)
            return rows

        raise serializers.ValidationError({"format": "不支持的导入格式。"})

    def _build_energy_data_instances(self, rows):
        devices = Device.objects.select_related("energy_type").all()
        device_by_pk = {str(device.pk): device for device in devices}
        device_by_code = {str(device.device_id): device for device in devices}

        energy_types = EnergyType.objects.all()
        energy_type_by_pk = {str(energy_type.pk): energy_type for energy_type in energy_types}
        energy_type_by_code = {
            str(energy_type.code).upper(): energy_type for energy_type in energy_types
        }

        instances = []
        for index, raw_row in enumerate(rows, start=1):
            if not isinstance(raw_row, dict):
                raise serializers.ValidationError({"row": f"第 {index} 行数据必须是对象。"})
            row = _normalize_row(raw_row)

            device_raw = _pick_value(
                row,
                ["device", "device_id", "device_pk", "meter_id", "nmi_id"],
            )
            if device_raw in (None, ""):
                raise serializers.ValidationError({"row": f"第 {index} 行缺少设备标识。"})
            device_key = str(device_raw).strip()
            device = device_by_pk.get(device_key) or device_by_code.get(device_key)
            if device is None:
                raise serializers.ValidationError({"row": f"第 {index} 行设备不存在: {device_raw}"})

            energy_raw = _pick_value(
                row,
                ["energy_type", "energy_type_id", "energy_type_code", "type", "type_code"],
            )
            if energy_raw in (None, ""):
                energy_type = device.energy_type
            else:
                energy_key = str(energy_raw).strip()
                energy_type = energy_type_by_pk.get(energy_key) or energy_type_by_code.get(
                    energy_key.upper()
                )
                if energy_type is None:
                    raise serializers.ValidationError({"row": f"第 {index} 行能源类型不存在: {energy_raw}"})
            if energy_type.id != device.energy_type_id:
                raise serializers.ValidationError(
                    {"row": f"第 {index} 行能源类型与设备绑定类型不一致。"}
                )

            timestamp_raw = _pick_value(row, ["timestamp", "time", "datetime", "record_time"])
            timestamp_value = _as_aware_datetime(timestamp_raw)
            if timestamp_value is None:
                raise serializers.ValidationError({"row": f"第 {index} 行时间戳格式错误。"})

            value_raw = _pick_value(row, ["value", "reading", "usage", "consumption"])
            value = _parse_decimal(value_raw, "value", required=True)
            voltage = _parse_decimal(_pick_value(row, ["voltage"]), "voltage")
            current = _parse_decimal(_pick_value(row, ["current"]), "current")
            power = _parse_decimal(_pick_value(row, ["power"]), "power")
            flow_rate = _parse_decimal(_pick_value(row, ["flow_rate", "flow"]), "flow_rate")

            instances.append(
                EnergyData(
                    device=device,
                    energy_type=energy_type,
                    timestamp=timestamp_value,
                    value=value,
                    voltage=voltage,
                    current=current,
                    power=power,
                    flow_rate=flow_rate,
                )
            )

        return instances

    def _export_excel(self, rows):
        try:
            import openpyxl
        except ImportError as exc:
            raise serializers.ValidationError({"format": "缺少 openpyxl 依赖，无法导出 Excel。"}) from exc

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "energy_data"
        headers = [
            "id",
            "device",
            "device_code",
            "device_name",
            "energy_type",
            "timestamp",
            "value",
            "voltage",
            "current",
            "power",
            "flow_rate",
        ]
        sheet.append(headers)

        for item in rows:
            sheet.append(
                [
                    item.id,
                    item.device_id,
                    item.device.device_id,
                    item.device.name,
                    item.energy_type.code,
                    item.timestamp.isoformat(),
                    str(item.value),
                    str(item.voltage) if item.voltage is not None else "",
                    str(item.current) if item.current is not None else "",
                    str(item.power) if item.power is not None else "",
                    str(item.flow_rate) if item.flow_rate is not None else "",
                ]
            )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="energy-data.xlsx"'
        return response

    def _export_pdf(self, rows):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
        except ImportError as exc:
            raise serializers.ValidationError({"format": "缺少 reportlab 依赖，无法导出 PDF。"}) from exc

        output = io.BytesIO()
        pdf = canvas.Canvas(output, pagesize=A4)
        page_width, page_height = A4

        y = page_height - 32
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(32, y, "Energy Data Export")
        y -= 24
        pdf.setFont("Helvetica", 9)
        pdf.drawString(32, y, "id | device_code | energy_type | timestamp | value")
        y -= 16

        for item in rows:
            line = (
                f"{item.id} | {item.device.device_id} | {item.energy_type.code} | "
                f"{item.timestamp.strftime('%Y-%m-%d %H:%M:%S')} | {item.value}"
            )
            pdf.drawString(32, y, line[:140])
            y -= 14
            if y < 36:
                pdf.showPage()
                y = page_height - 32
                pdf.setFont("Helvetica", 9)

        pdf.save()
        output.seek(0)

        response = HttpResponse(output.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="energy-data.pdf"'
        return response


@extend_schema_view(
    list=extend_schema(summary="查询能耗统计数据列表"),
)
class EnergyStatisticsViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    """能耗统计数据只读接口。"""

    queryset = EnergyStatistics.objects.select_related("device", "energy_type").all().order_by(
        "-period_date",
        "-id",
    )
    serializer_class = EnergyStatisticsSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["device__device_id", "device__name", "energy_type__code"]
    ordering_fields = ["id", "period_date", "period_type", "total_value", "created_at"]
    ordering = ["-period_date", "-id"]
    http_method_names = ["get", "head", "options"]

    def get_queryset(self):
        queryset = super().get_queryset()

        device_id = self.request.query_params.get("device_id")
        if device_id:
            if str(device_id).isdigit():
                queryset = queryset.filter(device_id=int(device_id))
            else:
                queryset = queryset.filter(device__device_id=device_id)

        energy_type = self.request.query_params.get("energy_type")
        if energy_type:
            if str(energy_type).isdigit():
                queryset = queryset.filter(energy_type_id=int(energy_type))
            else:
                queryset = queryset.filter(energy_type__code__iexact=energy_type)

        period_type = self.request.query_params.get("period_type")
        if period_type:
            queryset = queryset.filter(period_type=period_type)

        start_date = self.request.query_params.get("start_date")
        if start_date:
            parsed_start_date = parse_date(str(start_date))
            if parsed_start_date:
                queryset = queryset.filter(period_date__gte=parsed_start_date)

        end_date = self.request.query_params.get("end_date")
        if end_date:
            parsed_end_date = parse_date(str(end_date))
            if parsed_end_date:
                queryset = queryset.filter(period_date__lte=parsed_end_date)

        return queryset
